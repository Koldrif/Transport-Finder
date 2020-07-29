import xlwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from src.Styles import Styles_Excel

imgLogo = Image('.\src\Images\Logo.png')
imgLogo.width = 160
imgLogo.height = 50
save_file_name = 'openPyxl.xlsx'

def as_text(value):
    if value is None:
        return ""
    return str(value)

def old_format(inn, database):
    inn = str(inn)
    filename = ''
    data = database.get_data(
        'vin',                           # 0
        'srm',                           # 1
        'ownership',                     # 2       
        'end_of_ownership',              # 3                
        'registered_at_o',               # 4            
        'license_number',                # 5            
        'brand',                         # 6    
        'number_of_cat_reg',             # 7                
        'owner_from_cat_reg',            # 8                
        'model_from_cat_reg',            # 9                
        'atp',                           # 10
        'category',                      # 11    
        'type',                          # 12
        'ogrn',                          # 13
        'inspect_start',                 # 14            
        'form_of_holding_inspect',       # 15                   
        'purpose_of_inspect',            # 16               
        'other_reason_of_inspect',       # 17                    
        'company',                       # 18   
        'inspect_duration',              # 19
        inn=inn                          
        )
    recommendation = ''
    file_output = Workbook()
    sheet1 = file_output.active

    sheet1.merge_cells('A1:B1')
    sheet1.merge_cells('C1:I1')
    sheet1.merge_cells('A2:I2')
    sheet1.merge_cells('A3:E3')
    sheet1.merge_cells('F3:I3')
    sheet1.merge_cells('F4:I4')
    sheet1.merge_cells('F5:I5')
    sheet1.merge_cells('F6:I6')
    sheet1.merge_cells('F7:I7')
    sheet1.merge_cells('B4:E4')
    sheet1.merge_cells('B5:E5')
    sheet1.merge_cells('B6:E6')
    sheet1.merge_cells('B7:E7')
    sheet1.merge_cells('B8:E8')
    sheet1.merge_cells('A9:I9')
    sheet1.merge_cells('A10:I10')
    sheet1.merge_cells('A11:I11')
    sheet1.merge_cells('A12:I12')
    sheet1.merge_cells('A13:I13')
    sheet1.merge_cells('A14:I14')
    sheet1.merge_cells('A15:I15')
    sheet1.merge_cells('A16:I16')
    sheet1.merge_cells('A17:I17')
    sheet1.merge_cells('A18:I18')
    sheet1.merge_cells('A19:I19')
    sheet1.merge_cells('A20:I20')
    sheet1.merge_cells('A21:I21')
    sheet1.merge_cells('A22:E22')
    sheet1.merge_cells('A23:E23')
    sheet1.merge_cells('A26:E26')
    sheet1.merge_cells('A27:E27')
    sheet1.merge_cells('A31:C32')
    sheet1.merge_cells('D31:H32')
    sheet1.merge_cells('I31:N32')
    sheet1.add_image(imgLogo, 'A1')
    sheet1['A1'].style = Styles_Excel.Style_1
    sheet1['B1'].style = Styles_Excel.Style_1
    sheet1['C1'].style = Styles_Excel.Style_1
    sheet1['D1'].style = Styles_Excel.Style_1
    sheet1['E1'].style = Styles_Excel.Style_1
    sheet1['F1'].style = Styles_Excel.Style_1
    sheet1['G1'].style = Styles_Excel.Style_1
    sheet1['H1'].style = Styles_Excel.Style_1
    sheet1['I1'].style = Styles_Excel.Style_1
    #sheet1['A2'].style = Header_info_style
    #sheet1['A3'].style = Header_info_style
    #sheet1['A4'].style = Header_info_style
    #sheet1['A5'].style = Header_info_style
    # sheet1['B2'].style = Header_info_style
    # sheet1['B3'].style = Header_info_style
    # sheet1['B4'].style = Header_info_style
    # sheet1['B5'].style = Header_info_style
    # sheet1['D2'].style = Prosecutor_check_style
    # sheet1['D3'].style = Prosecutor_check_style
    # sheet1['D4'].style = Prosecutor_check_style
    # sheet1['D5'].style = Prosecutor_check_style
    # sheet1['E3'].style = Prosecutor_info_check_style
    # sheet1['E4'].style = Prosecutor_info_check_style
    # sheet1['E5'].style = Prosecutor_info_check_style
    #sheet1['G13'].style = Header_info_style

    sheet1.title = 'Company and vehicle information'
    sheet1['C1'] = 'ЗАЩИТА ИНТЕРЕСОВ БИЗНЕСА ПО ВСЕЙ РОССИИ'
    sheet1['A3'] = 'Аналитическая справка по ИНН' #TODO добавить контакты, обновить стиль под файл из watsup
    sheet1['A3'].style = Styles_Excel.Style_1_Center
    sheet1['B3'].style = Styles_Excel.Style_1_Center
    sheet1['C3'].style = Styles_Excel.Style_1_Center
    sheet1['D3'].style = Styles_Excel.Style_1_Center
    sheet1['E3'].style = Styles_Excel.Style_1_Center
    sheet1['A4'] = 'Для Компании'
    sheet1['A4'].style = Styles_Excel.Style_1_10px
    sheet1['C4'].style = Styles_Excel.Style_1_10px
    sheet1['D4'].style = Styles_Excel.Style_1_10px
    sheet1['E4'].style = Styles_Excel.Style_1_10px
    sheet1['B4'] = data[0][18]
    sheet1['B4'].style = Styles_Excel.Style_1_10px
    sheet1['C4'].style = Styles_Excel.Style_1_10px
    sheet1['D4'].style = Styles_Excel.Style_1_10px
    sheet1['E4'].style = Styles_Excel.Style_1_10px
    sheet1['A5'] = 'ИНН'
    sheet1['A5'].style = Styles_Excel.Style_1_10px
    sheet1['C5'].style = Styles_Excel.Style_1_10px
    sheet1['D5'].style = Styles_Excel.Style_1_10px
    sheet1['E5'].style = Styles_Excel.Style_1_10px
    sheet1['B5'] = inn
    sheet1['B5'].style = Styles_Excel.Style_1_10px
    sheet1['C5'].style = Styles_Excel.Style_1_10px
    sheet1['D5'].style = Styles_Excel.Style_1_10px
    sheet1['E5'].style = Styles_Excel.Style_1_10px
    sheet1['A6'] = 'ОГРН'
    sheet1['A6'].style = Styles_Excel.Style_1_10px
    sheet1['C6'].style = Styles_Excel.Style_1_10px
    sheet1['D6'].style = Styles_Excel.Style_1_10px
    sheet1['E6'].style = Styles_Excel.Style_1_10px
    sheet1['B6'] = data[0][13]
    sheet1['B6'].style = Styles_Excel.Style_1_10px
    sheet1['C6'].style = Styles_Excel.Style_1_10px
    sheet1['D6'].style = Styles_Excel.Style_1_10px
    sheet1['E6'].style = Styles_Excel.Style_1_10px
    sheet1['A7'] = '№ Лицензии'
    sheet1['A7'].style = Styles_Excel.Style_1_10px
    sheet1['C7'].style = Styles_Excel.Style_1_10px
    sheet1['D7'].style = Styles_Excel.Style_1_10px
    sheet1['E7'].style = Styles_Excel.Style_1_10px
    sheet1['B7'] = data[0][5]
    sheet1['B7'].style = Styles_Excel.Style_1_10px
    sheet1['C7'].style = Styles_Excel.Style_1_10px
    sheet1['D7'].style = Styles_Excel.Style_1_10px
    sheet1['E7'].style = Styles_Excel.Style_1_10px
    sheet1['A8'] = 'Дата Лицензии'
    sheet1['A8'].style = Styles_Excel.Style_1_10px
    sheet1['C8'].style = Styles_Excel.Style_1_10px
    sheet1['D8'].style = Styles_Excel.Style_1_10px
    sheet1['E8'].style = Styles_Excel.Style_1_10px
    sheet1['B8'] = data[0][4]
    sheet1['B8'].style = Styles_Excel.Style_1_10px
    sheet1['C8'].style = Styles_Excel.Style_1_10px
    sheet1['D8'].style = Styles_Excel.Style_1_10px
    sheet1['E8'].style = Styles_Excel.Style_1_10px
    sheet1['F3'] = '117342, г. Москва, ул Бутлерова, дом 17, БЦ Нео Гео'
    sheet1['F3'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['G3'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['H3'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['I3'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['F4'] = 'ПИШИТЕ WHATSAPP +7(926)862-02-09'
    sheet1['F4'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['G4'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['H4'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['I4'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['F5'] = 'ЗВОНИТЕ 8 (495) 021-22-28'
    sheet1['F5'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['G5'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['H5'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['I5'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['F6'] = 'ПИШИТЕ YOURSPEC@YOURSPEC.NET'
    sheet1['F6'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['G6'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['H6'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['I6'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['F7'] = 'yourspec.net'
    sheet1['F7'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['G7'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['H7'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['I7'].style = Styles_Excel.Style_1_ligh_Green_10px_center
    sheet1['A10'] = 'ПАМЯТКА ПЕРЕВОЗЧИКУ:'
    sheet1['A10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['B10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['C10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['D10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['E10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['F10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['G10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['H10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['I10'].style = Styles_Excel.Style_1_Red_Center
    sheet1['A11'] = '1. Не требуется делать Категорирование, Оценку уязвимости, План безопасности транспортного средства:'
    sheet1['A12'] = '- Если Вы перевозите учащихся от места проживания к месту обучения и обратно на безвозмездной основе.'
    sheet1['A13'] = '- Если Вы осуществляете перевозки в целях оказания ритуальных услуг.'
    sheet1['A14'] = '- Не нужно делать на прицепы, полуприцепы, используемые для перевозки опасных грузов.'
    sheet1['A15'] = '2. Перевозчик делает Категорирование, Оценку уязвимости, План безопасности транспортного средства - Один раз и навсегда.'
    sheet1['A16'] = '3. Переделывать документы требуется только если сменился собственник транспортного средства.'
    sheet1['A17'] = '4. Группировка транспортных средств - это объединение одинаковых транспортных средств по модельному ряду в один документ.'
    sheet1['A18'] = '5. Количеству групп Оценок уязвимости должно совпадать с количеством групп по Планам безопасности транспортных средств.'
    sheet1['A19'] = '6. Исправить данные в Реестре требуется когда Росавтодор ввел некорректные данные.'
    sheet1['A20'] = '7. Обязательно исключать из Реестра транспорт который продан!'
    sheet1['A22'] = 'ПЛАНОВАЯ ПРОВЕРКА БУДЕТ ПРОВЕДЕНА'
    sheet1['A22'].style = Styles_Excel.Style_1_Red_Center
    sheet1['B22'].style = Styles_Excel.Style_1_Red_Center
    sheet1['C22'].style = Styles_Excel.Style_1_Red_Center
    sheet1['D22'].style = Styles_Excel.Style_1_Red_Center
    sheet1['E22'].style = Styles_Excel.Style_1_Red_Center
    sheet1['A23'] = 'Согласно ст. 27 Постановление Правительства РФ от 27.02.2019 N 195 "О лицензировании деятельности по перевозкам пассажиров и иных лиц автобусами"'
    sheet1['A24'] = 'Дата проведения проверки'
    sheet1['A24'].style = Styles_Excel.Bold_Font_subhead
    sheet1['B24'] = 'с'
    sheet1['B24'].style = Styles_Excel.Bold_Font_subhead
    sheet1['C24'] =  data[0][14] #? Как форматировать дату
    sheet1['C24'].style = Styles_Excel.Bold_Font_subhead
    sheet1['D24'] =  'до' 
    sheet1['D24'].style = Styles_Excel.Bold_Font_subhead
    sheet1['E24'] =  '' #!
    sheet1['E24'].style = Styles_Excel.Bold_Font_subhead
    sheet1['A26'] =  'ПРОКУРОРСКАЯ ПРОВЕРКА БУДЕТ ПРОВЕДЕНА' 
    sheet1['A26'].style = Styles_Excel.Style_1_Red_Center
    sheet1['B26'].style = Styles_Excel.Style_1_Red_Center
    sheet1['C26'].style = Styles_Excel.Style_1_Red_Center
    sheet1['D26'].style = Styles_Excel.Style_1_Red_Center
    sheet1['E26'].style = Styles_Excel.Style_1_Red_Center
    sheet1['A27'] =   data[0][16]
    sheet1['A28'] =   'Месяц проведения проверки'
    sheet1['A28'].style = Styles_Excel.Bold_Font_subhead 
    sheet1['A29'] =   data[0][14]
    sheet1['B28'] =   'Рабочих дней'
    sheet1['B28'].style = Styles_Excel.Bold_Font_subhead
    sheet1['B29'] =   data[0][14] #? нужно допилить
    sheet1['C28'] =   'Рабочих часов'
    sheet1['C28'].style = Styles_Excel.Bold_Font_subhead
    sheet1['C29'] =   data[0][19]
    sheet1['D28'] =   'Форма проведения проверки'
    sheet1['D28'].style = Styles_Excel.Bold_Font_subhead
    sheet1['D29'] =   data[0][15]
    sheet1['E28'] =   'Другие причины проверки'
    sheet1['E28'].style = Styles_Excel.Bold_Font_subhead
    sheet1['E29'] =   data[0][17]
    sheet1['A31'] =   'ТРЕБУЕТСЯ СДЕЛАТЬ'
    sheet1['A31'].style = Styles_Excel.Style_1_Red_Center
    sheet1['A32'].style = Styles_Excel.Style_1_Red_Center
    sheet1['B31'].style = Styles_Excel.Style_1_Red_Center
    sheet1['B32'].style = Styles_Excel.Style_1_Red_Center
    sheet1['C31'].style = Styles_Excel.Style_1_Red_Center
    sheet1['C32'].style = Styles_Excel.Style_1_Red_Center
    sheet1['D31'] =   'НАЙДЕН ТРАНСПОРТ В РЕЕСТРЕ ЛИЦЕНЗИЙ'
    sheet1['D31'].style = Styles_Excel.Style_1_Center
    sheet1['D32'].style = Styles_Excel.Style_1_Center
    sheet1['E31'].style = Styles_Excel.Style_1_Center
    sheet1['E32'].style = Styles_Excel.Style_1_Center
    sheet1['F31'].style = Styles_Excel.Style_1_Center
    sheet1['F32'].style = Styles_Excel.Style_1_Center
    sheet1['G31'].style = Styles_Excel.Style_1_Center
    sheet1['G32'].style = Styles_Excel.Style_1_Center
    sheet1['H31'].style = Styles_Excel.Style_1_Center
    sheet1['H32'].style = Styles_Excel.Style_1_Center
    sheet1['I31'] =   'НАЙДЕН ТРАНСПОРТ В РЕЕСТРЕ КАТЕГОРИРОВАНИЯ'
    sheet1['I31'].style = Styles_Excel.Style_1_Center
    sheet1['I32'].style = Styles_Excel.Style_1_Center
    sheet1['J31'].style = Styles_Excel.Style_1_Center
    sheet1['J32'].style = Styles_Excel.Style_1_Center
    sheet1['K31'].style = Styles_Excel.Style_1_Center
    sheet1['K32'].style = Styles_Excel.Style_1_Center
    sheet1['L31'].style = Styles_Excel.Style_1_Center
    sheet1['L32'].style = Styles_Excel.Style_1_Center
    sheet1['M31'].style = Styles_Excel.Style_1_Center
    sheet1['M32'].style = Styles_Excel.Style_1_Center
    sheet1['N31'].style = Styles_Excel.Style_1_Center
    sheet1['N32'].style = Styles_Excel.Style_1_Center
    sheet1['A33'] =   'Категорирование'
    sheet1['A33'].style = Styles_Excel.Style_1_Red_10px
    sheet1['B33'] =   'Оценка уязвимости'
    sheet1['B33'].style = Styles_Excel.Style_1_Red_10px
    sheet1['C33'] =   'План безопасности'
    sheet1['C33'].style = Styles_Excel.Style_1_Red_10px
    sheet1['D33'] =   'Модель по ПТС'
    sheet1['D33'].style = Styles_Excel.Style_1_10px_center
    sheet1['E33'] =   'Гос. рег. номер'
    sheet1['E33'].style = Styles_Excel.Style_1_10px_center
    sheet1['F33'] =   'VIN'
    sheet1['F33'].style = Styles_Excel.Style_1_10px_center
    sheet1['G33'] =   'Право владения'
    sheet1['G33'].style = Styles_Excel.Style_1_10px_center
    sheet1['H33'] =   'Статус'
    sheet1['H33'].style = Styles_Excel.Style_1_10px_center
    sheet1['I33'] =   '№ реестра'
    sheet1['I33'].style = Styles_Excel.Style_1_10px_center
    sheet1['J33'] =   'АТП №'
    sheet1['J33'].style = Styles_Excel.Style_1_10px_center
    sheet1['K33'] =   'Тип транспорта'
    sheet1['K33'].style = Styles_Excel.Style_1_10px_center
    sheet1['L33'] =   'Модель из Реестра'
    sheet1['L33'].style = Styles_Excel.Style_1_10px_center
    sheet1['M33'] =   'Собственник'
    sheet1['M33'].style = Styles_Excel.Style_1_10px_center
    sheet1['N33'] =   'Категория транспорта'
    sheet1['N33'].style = Styles_Excel.Style_1_10px_center
    for i in range(len(data)):
        #sheet1.cell(i+33, 3, data[i][]) #TODO Модель по АТП
        sheet1.cell(i+34, 5, data[i][1]) # Гос рег номер
        sheet1.cell(i+34, 6, data[i][0])
        sheet1.cell(i+34, 7, data[i][2])
        sheet1.cell(i+34, 8, data[i][3])
        #sheet1.cell(i+33+1, 8, data[i][10]) #! Тут нужно имя файла
        sheet1.cell(i+34, 10, data[i][10]) # АТП
        sheet1.cell(i+34, 11, data[i][12])
        sheet1.cell(i+34, 12, data[i][9])
        sheet1.cell(i+34, 13, data[i][8])
        sheet1.cell(i+34, 14, data[i][11])
        if ((data[i][10] == '' or data[i][10] == 'Н/Д' or data[i][12] == '' or data[i][12] == 'Н/Д' or data[i][9] == '' or  data[i][9] == 'Н/Д' or data[i][8] == '' or  data[i][8] == 'Н/Д' or data[i][11] == '' or data[i][11] == 'Н/Д')):
            sheet1.cell(i+34, 1, 'Категорировать')
            sheet1.cell(i+34, 1).style = Styles_Excel.Attention_Style
            sheet1.cell(i+34, 2, 'Требуется оценка уязвимости')
            sheet1.cell(i+34, 2).style = Styles_Excel.Attention_Style
            sheet1.cell(i+34, 3, 'Требуется план безопасности')
            sheet1.cell(i+34, 3).style = Styles_Excel.Attention_Style

        else:
            sheet1.cell(i+34, 1, 'Не требуется') #? Под вопросом, нужно еще несколько условий, это просто для заполнения
            sheet1.cell(i+34, 1).style = Styles_Excel.Ok_Style                       

    dims = {}
    for row in sheet1.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        sheet1.column_dimensions[chr(col+64)].width = value + 5


    sheet1.column_dimensions['A'].width = 30
    sheet1.column_dimensions['B'].width = 30
    sheet1.column_dimensions['C'].width = 30
    sheet1.column_dimensions['D'].width = 30
    sheet1.column_dimensions['E'].width = 30
    sheet1.column_dimensions['F'].width = 30
    sheet1.column_dimensions['I'].width = 30
    sheet1.column_dimensions['N'].width = 30
    sheet1.row_dimensions[1].height = 40

    file_output.save(save_file_name)
    return save_file_name, recommendation

    