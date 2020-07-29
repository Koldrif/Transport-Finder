from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Alignment


border = Side(style='thin', color="000000")

#Style_1 - первые четыре клетки по памятке А:11 - А:14

Bold_Font_header = NamedStyle(name='Bold_Font_header', 
                        border=Border(left=border, 
                                      top=border, 
                                      right=border, 
                                      bottom=border),
                         font=Font(name='Arial', bold=True, size=14, vertAlign='baseline'),
                         alignment=Alignment(horizontal='center')
                              ) # , vertical='center'))

Bold_Font_subhead = NamedStyle(name='Bold_Font_subhead', 
                    #    border=Border(left=border, 
                    #                  top=border, 
                    #                  right=border, 
                    #                  bottom=border),
                       font=Font(name='Arial', bold=True, size=10 ))

Attention_Style = NamedStyle(name='Attention_Style', 
                    fill=PatternFill('solid', 
                                     fgColor='ff6666'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=False, size=10 )) 

Ok_Style = NamedStyle(name='Ok_Style', 
                    font=Font(name='Arial', bold=False, size=10 ))                                          

Style_1 = NamedStyle(name='Style_1', 
                    fill=PatternFill('solid', 
                                     fgColor='ccffff'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=14 ))

Style_1_10px = NamedStyle(name='Style_1_10px', 
                    fill=PatternFill('solid', 
                                     fgColor='ccffff'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=10 ))

Style_1_Center = NamedStyle(name='Style_1_Center', 
                    fill=PatternFill('solid', 
                                     fgColor='ccffff'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=14 ),
                    alignment=Alignment(horizontal='center',
                    vertical='center'))

Style_1_Red_Center = NamedStyle(name='Style_1_Red_Center', 
                    fill=PatternFill('solid', 
                                     fgColor='ff6666'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=14 ),
                    alignment=Alignment(horizontal='center',
                    vertical='center'))

Style_1_10px_center = NamedStyle(name='Style_1_10px_center', 
                    fill=PatternFill('solid', 
                                     fgColor='ccffff'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=10 ),
                    alignment=Alignment(horizontal='center',
                    vertical='center'))

Style_1_ligh_Green_10px_center = NamedStyle(name='Style_1_ligh_Green_10px_center', 
                    fill=PatternFill('solid', 
                                     fgColor='BFFFDB'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=10 ),
                    alignment=Alignment(horizontal='center',
                    vertical='center'))

Style_1_Red_10px = NamedStyle(name='Style_1_Red_10px', 
                    fill=PatternFill('solid', 
                                     fgColor='ff6666'),
                    border=Border(left=border, 
                                  top=border, 
                                  right=border, 
                                  bottom=border),
                    font=Font(name='Arial', bold=True, size=10 ),
                    alignment=Alignment(horizontal='center',
                    vertical='center'))

Style_2 = NamedStyle(name='Style_2', 
                              fill=PatternFill('solid', 
                                               fgColor='00f4cccc'), 

                              border=Border(left=border, 
                                            top=border, 
                                            right=border, 
                                            bottom=border))
